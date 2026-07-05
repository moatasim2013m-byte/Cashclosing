#!/usr/bin/env python3
"""Audit a "kids master list" export and prepare it for the Birthdays sheet.

Usage:
    python3 scripts/check_kids_list.py <kids_list.xlsx> [conversations.csv]

What it does:
    1. Validates every phone number (Jordanian mobile format, junk detection).
    2. Reports birthday coverage (Confirmed vs Missing).
    3. Optionally cross-checks numbers against a conversations CSV export
       (the "Peekaboo Conversations" sheet) to spot recently active customers.
    4. Writes birthdays_import.csv with kids that have BOTH a confirmed
       birthday and a valid phone — ready to paste into the "Birthdays" tab
       used by the app's Birthday Reminders feature.

Expected xlsx columns (sheet "All Kids" or first sheet):
    First Name | Last Name | Mobile | Date Of Birth | Birthday Status | Visit Count
Dates are read as MM/DD/YYYY (POS export format) or ISO.
"""

import csv
import io
import re
import sys
from collections import Counter
from datetime import date

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl is required: pip install openpyxl")

COUNTRY_CODE = "962"


def norm(p):
    if p is None:
        return ""
    d = re.sub(r"\D", "", str(p))
    if not d:
        return ""
    if d.startswith("00"):
        d = d[2:]
    if d.startswith("0") and len(d) == 10:
        d = COUNTRY_CODE + d[1:]
    elif len(d) == 9 and d.startswith("7"):
        d = COUNTRY_CODE + d
    return d


def is_valid_jo(d):
    return len(d) == 12 and d.startswith("9627") and d[3:5] in ("77", "78", "79")


def is_valid_intl(d):
    return 11 <= len(d) <= 14 and not d.startswith("0")


def fix_dob(d):
    m = re.match(r"^(\d{1,2})/(\d{1,2})/(\d{4})$", str(d).strip())
    if m:  # MM/DD/YYYY (POS export)
        return f"{m.group(3)}-{int(m.group(1)):02d}-{int(m.group(2)):02d}"
    m = re.match(r"^(\d{4})-(\d{1,2})-(\d{1,2})", str(d).strip())
    if m:
        return f"{m.group(1)}-{int(m.group(2)):02d}-{int(m.group(3)):02d}"
    return None


def main():
    if len(sys.argv) < 2:
        sys.exit(__doc__)
    xlsx_path = sys.argv[1]
    convo_path = sys.argv[2] if len(sys.argv) > 2 else None

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["All Kids"] if "All Kids" in wb.sheetnames else wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))[1:]

    kids = []
    for r in rows:
        first, last, mobile, dob, status, visits = (list(r) + [None] * 6)[:6]
        candidates = [norm(mobile), norm(last)]  # a 2nd number sometimes sits in Last Name
        phone = next((c for c in candidates if is_valid_jo(c)), "") or next(
            (c for c in candidates if is_valid_intl(c)), ""
        )
        kids.append(
            {
                "first": str(first or "").strip(),
                "last": str(last or "").strip(),
                "raw_mobile": mobile,
                "phone": phone,
                "dob": dob,
                "status": str(status or ""),
                "visits": visits or 0,
            }
        )

    total = len(kids)
    with_phone = [k for k in kids if k["phone"]]
    junk = [k for k in kids if k["raw_mobile"] not in (None, "") and not k["phone"]]
    confirmed = [k for k in kids if k["status"] == "Confirmed"]
    ready = [k for k in confirmed if k["phone"] and fix_dob(k["dob"])]

    print(f"KIDS LIST: {total} rows")
    print(f"  valid phone:        {len(with_phone)} ({len(with_phone)/total:.0%})")
    print(f"  junk/invalid phone: {len(junk)}")
    if junk:
        print("    samples:", Counter(str(k["raw_mobile"]) for k in junk).most_common(5))
    print(f"  unique phones:      {len(set(k['phone'] for k in with_phone))}")
    print(f"  confirmed birthday: {len(confirmed)} ({len(confirmed)/total:.0%})")
    print(f"  READY (birthday + phone): {len(ready)}")

    if convo_path:
        convo_rows = list(csv.reader(open(convo_path, encoding="utf-8-sig")))
        idx = next(i for i, h in enumerate(convo_rows[0]) if "phone" in h.lower())
        convo_phones = {
            norm(r[idx]) for r in convo_rows[1:] if len(r) > idx and is_valid_intl(norm(r[idx]))
        }
        master_phones = {k["phone"] for k in with_phone}
        overlap = master_phones & convo_phones
        print(f"\nCROSS-CHECK vs {convo_path}:")
        print(f"  conversation numbers: {len(convo_phones)}")
        print(f"  also in kids list:    {len(overlap)}")
        print(f"  NOT in kids list:     {len(convo_phones - overlap)} (new customers, no kid record)")

    out = "birthdays_import.csv"
    with open(out, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["Kid Name", "Birth Date", "Parent Name", "Phone", "Last Wished"])
        for k in sorted(ready, key=lambda k: fix_dob(k["dob"])[5:]):
            name = f"{k['first']} {k['last']}".strip().strip("];=").strip()
            w.writerow([name, fix_dob(k["dob"]), "", k["phone"], ""])
    print(f"\nWrote {len(ready)} rows to {out} (paste into the 'Birthdays' sheet tab)")

    today = date.today()
    soon = []
    for k in ready:
        iso = fix_dob(k["dob"])
        m, d = int(iso[5:7]), int(iso[8:10])
        for yr in (today.year, today.year + 1):
            try:
                nb = date(yr, m, d)
            except ValueError:
                nb = date(yr, 2, 28)
            if nb >= today:
                break
        du = (nb - today).days
        if du <= 15:
            soon.append((du, k, iso))
    soon.sort(key=lambda t: t[0])
    print(f"\nBirthdays in the next 15 days: {len(soon)}")
    for du, k, iso in soon:
        print(f"  in {du:2d} days — {k['first']} {k['last']} (born {iso}) — {k['phone']}")


if __name__ == "__main__":
    main()
